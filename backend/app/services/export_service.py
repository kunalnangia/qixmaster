"""
Test Case Export Service
Handles exporting test cases to various formats (PDF, Excel, CSV)
"""
import io
import csv
from datetime import datetime
from typing import List, Dict, Any, Optional
import pandas as pd
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import json


class TestCaseExportService:
    """Service for exporting test cases to various formats"""
    
    @staticmethod
    def export_to_csv(test_cases: List[Dict[str, Any]], project_name: str = "Test Cases") -> io.StringIO:
        """Export test cases to CSV format"""
        output = io.StringIO()
        
        if not test_cases:
            writer = csv.writer(output)
            writer.writerow(["No test cases to export"])
            output.seek(0)
            return output
        
        # Define CSV headers
        headers = [
            "Title", "Description", "Test Type", "Priority", "Status", 
            "Expected Result", "Tags", "Prerequisites", "Steps", "Created By"
        ]
        
        writer = csv.writer(output)
        writer.writerow(headers)
        
        for tc in test_cases:
            # Format steps as numbered list
            steps_text = ""
            if tc.get("test_steps"):
                steps = []
                for i, step in enumerate(tc.get("test_steps", []), 1):
                    steps.append(f"{i}. {step.get('description', '')} -> {step.get('expected_result', '')}")
                steps_text = " | ".join(steps)
            
            # Write test case row
            writer.writerow([
                tc.get("title", ""),
                tc.get("description", ""),
                tc.get("test_type", ""),
                tc.get("priority", ""),
                tc.get("status", ""),
                tc.get("expected_result", ""),
                ", ".join(tc.get("tags", [])),
                tc.get("preconditions", ""),
                steps_text,
                tc.get("created_by", "")
            ])
        
        output.seek(0)
        return output
    
    @staticmethod
    def export_to_excel(test_cases: List[Dict[str, Any]], project_name: str = "Test Cases") -> io.BytesIO:
        """Export test cases to Excel format"""
        output = io.BytesIO()
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        if ws is not None:
            ws.title = "Test Cases"
        
        # Define headers
        headers = [
            "ID", "Title", "Description", "Test Type", "Priority", "Status",
            "Expected Result", "Tags", "Prerequisites", "Steps", "Created By", "Created At"
        ]
        
        # Style for headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Add headers
        if ws is not None:
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                if cell is not None:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
        
        # Add test case data
        if ws is not None:
            for row, tc in enumerate(test_cases, 2):
                # Format steps as numbered list
                steps_text = ""
                if tc.get("test_steps"):
                    steps = []
                    for i, step in enumerate(tc.get("test_steps", []), 1):
                        steps.append(f"{i}. {step.get('description', '')} -> {step.get('expected_result', '')}")
                    steps_text = "\n".join(steps)
                
                # Add data to cells
                ws.cell(row=row, column=1, value=tc.get("id", ""))
                ws.cell(row=row, column=2, value=tc.get("title", ""))
                ws.cell(row=row, column=3, value=tc.get("description", ""))
                ws.cell(row=row, column=4, value=tc.get("test_type", ""))
                ws.cell(row=row, column=5, value=tc.get("priority", ""))
                ws.cell(row=row, column=6, value=tc.get("status", ""))
                ws.cell(row=row, column=7, value=tc.get("expected_result", ""))
                ws.cell(row=row, column=8, value=", ".join(tc.get("tags", [])))
                ws.cell(row=row, column=9, value=tc.get("preconditions", ""))
                ws.cell(row=row, column=10, value=steps_text)
                ws.cell(row=row, column=11, value=tc.get("created_by", ""))
                ws.cell(row=row, column=12, value=tc.get("created_at", ""))
        
        # Auto-adjust column widths
        if ws is not None:
            for col in range(1, len(headers) + 1):
                max_length = 0
                column = get_column_letter(col)
                column_cells = ws[column] if ws is not None else []
                for cell in column_cells:
                    try:
                        if cell.value is not None and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                if ws.column_dimensions is not None:
                    ws.column_dimensions[column].width = adjusted_width
        
        # Save to BytesIO
        try:
            wb.save(output)
        except Exception:
            pass  # Handle case where wb might be None
        output.seek(0)
        return output
    
    @staticmethod
    def export_to_pdf(test_cases: List[Dict[str, Any]], project_name: str = "Test Cases") -> io.BytesIO:
        """Export test cases to PDF format"""
        output = io.BytesIO()
        
        # Create PDF document
        doc = SimpleDocTemplate(output, pagesize=A4)
        story = []
        
        # Get styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1  # Center alignment
        )
        
        # Add title
        title = Paragraph(f"Test Cases Export - {project_name}", title_style)
        story.append(title)
        story.append(Spacer(1, 20))
        
        # Add export info
        export_info = Paragraph(
            f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}<br/>"
            f"Total Test Cases: {len(test_cases)}",
            styles['Normal']
        )
        story.append(export_info)
        story.append(Spacer(1, 20))
        
        # Add each test case
        for i, tc in enumerate(test_cases, 1):
            # Test case header
            tc_title = Paragraph(f"<b>{i}. {tc.get('title', 'Untitled Test Case')}</b>", styles['Heading2'])
            story.append(tc_title)
            
            # Test case details table
            data = [
                ['Field', 'Value'],
                ['ID', tc.get('id', '')],
                ['Description', tc.get('description', '')],
                ['Test Type', tc.get('test_type', '')],
                ['Priority', tc.get('priority', '')],
                ['Status', tc.get('status', '')],
                ['Expected Result', tc.get('expected_result', '')],
                ['Tags', ', '.join(tc.get('tags', []))],
                ['Prerequisites', tc.get('preconditions', '')],
            ]
            
            # Add steps if they exist
            if tc.get('test_steps'):
                steps_text = ""
                for j, step in enumerate(tc.get('test_steps', []), 1):
                    steps_text += f"{j}. {step.get('description', '')}\n   Expected: {step.get('expected_result', '')}\n"
                data.append(['Steps', steps_text])
            
            # Create table
            table = Table(data, colWidths=[2*inch, 4*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ]))
            
            story.append(table)
            story.append(Spacer(1, 20))
        
        # Build PDF
        doc.build(story)
        output.seek(0)
        return output
    
    @classmethod
    def get_export_filename(cls, project_name: str, export_format: str) -> str:
        """Generate appropriate filename for export"""
        # Sanitize project name for filename
        safe_name = "".join(c for c in project_name if c.isalnum() or c in (' ', '-', '_')).rstrip()
        safe_name = safe_name.replace(' ', '_')
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        return f"test_cases_{safe_name}_{timestamp}.{export_format}"